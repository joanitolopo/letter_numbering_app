from flask import Flask, render_template, request, redirect, url_for, flash, Response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import extract
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///numbering.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your-secret-key'
db = SQLAlchemy(app)

# Model Database
class Document(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sequence_number = db.Column(db.Integer, nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=True)
    officer_names = db.Column(db.String(500), nullable=False)
    activity = db.Column(db.String(200), nullable=False)
    document_number = db.Column(db.String(50), nullable=False, unique=True)

# Daftar nama pegawai

OFFICERS = [
    "Agustina S. Uly, A.Md.Keb",
    "Ester Y. Ie Kore, A.Md.Kep",
    "Corrina R. W. Lay, A.Md.Kep",
    "Kristian B. Dala Ngapa, A.Md.KG",
    "Jacob S. Radjah, A.Md.Kep",
    "Mariana Bau Bani",
    "Maria Y. Nau",
    "Aplonia Kana Wadu, S.Kep",
    "Jilian Frhans, S.KM",
    "Vesti E. Mata, S.Kep.,Ns",
    "Dewanti Z. Gani, A.Md.Kep",
    "Jhon K. Christian, A.Md.Farm",
    "Ina Irna M. Mangngi, A.Md.Keb",
    "Yudi K. H. R. Andung, A.Md.Gz",
    "Farida M. Lalang, A.Md.F",
    "Miati T. Bara Pa, A.Md.Kep",
    "Defiliance N. Kaseh, A.Md.Kes",
    "Debriana Pian, A.Md.KL",
    "Mario P. Rangga, A.Md.AK",
    "Dina Lena Kaho, A.Md.Gz",
    "Dewi S. Nubatonis, A.Md.KG",
    "Floriana W. Mema, A.Md.Gz",
    "Armi P. Boimau, A.Md.RMIK",
    "Bd. Patrisia Anna Maria, S.Tr.Keb",
    "Anita Mentari Rihi, A.Md.Keb",
    "Dennis M. Obidaka, A.Md.K",
    "Saratista H. Kale Te, A.Md.Kep",
    "Kurniati Balu Haba, A.Md.Keb",
    "Yanti Delfiana Radja, A.Md.Kep",
    "dr. Runiyuftari Lobo Huki",
    "dr. Christine Dupe",
    "Rio S. Bara, A.Md.Kep",
    "Oly A. Sabat, S.Kep.,Ns",
    "Sapta P. Adam, A.Md.Keb",
    "Dahlin Alexander Rohi, S.Kep.,Ns",
    "Citra Ades Bale Tadjo, A.Md.Kep",
    "Yusmiwati N. Selan, A.Md.Keb",
    "Lorensius, A.Md.Kep",
    "Salomy A. Ara Riwu, A.Md.Keb",
    "Pujiati, A.Md.Keb",
    "Chentil S. Petrus, A.Md.Keb",
    "Gaby M. Mira Manggi, S.KM",
    "Priska N. Djo, S.Kep.,Ns",
    "Maeyer N. S. E. Galla, A.Md.Keb",
    "Mayda Y. Br Singarimbun, A.Md.Keb",
    "Raden K. M. Lobo, A.Md.Keb",
    "Lesly Damayanti, A.Md.AK",
    "Indra Nawa, A.Md.KL",
    "Novita Doko, A.Md.Keb",
    "Hege Helena Djita, S.KM",
    "Berlian A. Pati, A.Md.Keb",
    "Ayu D. F. Br Sitepu, A.Md.Keb",
    "Alexs Lu Ndawa, A.Md.Kep",
    "Tri D. Hide Lilo, S.KM",
    "Gita Gewe, A.Md.Keb",
    "Febronia S Tameon, S.Gz",
    "Meriana Ratu Dubu, AMd Kep",
    "Yusni Lede, A.Md.Keb",
    "Lappa Migu, A.Md.Keb",
    "Elisabet Bani, S.Kep",
    "Sarche I. Loban, A.Md.Keb",
    "Jeenmarly N. Taga, A.Md.Kep",
    "Abdon M. Bani",
    "Yance Lodo Ga",
    "Johanis Kale Lay",
    "Luisa Mira Manngi",
    "Nofriance Lena Kaho",
    "Azarias Lena Kaho"
]

def month_to_roman(month):
    roman_months = {
        1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V', 6: 'VI',
        7: 'VII', 8: 'VIII', 9: 'IX', 10: 'X', 11: 'XI', 12: 'XII'
    }
    return roman_months[month]

def generate_document_number(sequence_number, date):
    year = date.strftime('%Y')
    month_roman = month_to_roman(date.month)
    sequence_str = f"{sequence_number:02d}"
    return f"800.1.11.1/{sequence_str}/UPTD PKM.B-SR/{month_roman}/{year}"

def is_sunday(date):
    return date.weekday() == 6

def check_date_overlap(start_date, end_date, exclude_id=None):
    """Cek apakah rentang tanggal tumpang tindih dengan dokumen lain."""
    existing_docs = Document.query.all()
    if exclude_id:
        existing_docs = [doc for doc in existing_docs if doc.id != exclude_id]
    
    for doc in existing_docs:
        doc_start = doc.start_date
        doc_end = doc.end_date if doc.end_date else doc.start_date
        if not (end_date < doc_start or start_date > doc_end):
            return True  # Ada tumpang tindih
    return False

def check_officer_conflict(start_date, end_date, officer_names, exclude_id=None):
    """Cek apakah petugas sudah terdaftar di dokumen lain dalam rentang."""
    date_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
    existing_docs = Document.query.all()
    if exclude_id:
        existing_docs = [doc for doc in existing_docs if doc.id != exclude_id]
    
    for date in date_range:
        for doc in existing_docs:
            doc_start = doc.start_date
            doc_end = doc.end_date if doc.end_date else doc.start_date
            if doc_start <= date <= doc_end:
                existing_officers = set(doc.officer_names.split(';'))
                for officer in officer_names:
                    if officer in existing_officers:
                        return True, officer, date
    return False, None, None

# def get_next_sequence_number(date, used_sequences):
#     year = date.year
#     month = date.month
#     key = (year, month)
    
#     # Ambil sequence_number tertinggi dari database
#     max_doc = db.session.query(Document).filter(
#         extract('year', Document.date) == year,
#         extract('month', Document.date) == month
#     ).order_by(Document.sequence_number.desc()).first()
    
#     # Tentukan sequence_number awal dari database
#     current_max = max_doc.sequence_number if max_doc else 0
    
#     # Perbarui dengan sequence_number yang sudah digunakan dalam transaksi
#     used_max = used_sequences.get(key, 0)
#     next_number = max(current_max, used_max) + 1
    
#     # Simpan sequence_number baru untuk tracking dalam transaksi
#     used_sequences[key] = next_number
#     print(f"get_next_sequence_number for {year}-{month}: current_max={current_max}, used_max={used_max}, returning {next_number}")
#     return next_number

# def resequence_documents(date):
#     year = date.year
#     month = date.month
#     docs = Document.query.filter(
#         extract('year', Document.date) == year,
#         extract('month', Document.date) == month
#     ).order_by(Document.date.asc(), Document.sequence_number.asc()).all()
#     for index, doc in enumerate(docs, start=1):
#         if doc.sequence_number != index:
#             old_sequence = doc.sequence_number
#             doc.sequence_number = index
#             doc.document_number = generate_document_number(index, doc.date)
#             print(f"Resequencing ID {doc.id}: old sequence={old_sequence}, new sequence={index}, new document_number={doc.document_number}")
#     db.session.commit()

@app.route('/', methods=['GET', 'POST'])
def index():
    selected_date = None
    if request.method == 'POST':
        date_str = request.form.get('filter_date')
        if date_str:
            try:
                selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                documents = Document.query.filter(
                    (Document.start_date <= selected_date) &
                    ((Document.end_date >= selected_date) | (Document.end_date == None))
                ).order_by(Document.start_date.asc(), Document.sequence_number.asc()).all()
            except ValueError:
                flash('Format tanggal tidak valid!', 'error')
                documents = Document.query.order_by(Document.start_date.asc(), Document.sequence_number.asc()).all()
        else:
            documents = Document.query.order_by(Document.start_date.asc(), Document.sequence_number.asc()).all()
    else:
        documents = Document.query.order_by(Document.start_date.asc(), Document.sequence_number.asc()).all()
    return render_template('index.html', documents=documents, selected_date=selected_date)

@app.route('/add', methods=['GET', 'POST'])
def add_document():
    if request.method == 'POST':
        try:
            start_date_str = request.form['start_date']
            end_date_str = request.form.get('end_date')
            sequence_number = request.form['sequence_number']
            document_number = request.form['document_number'].strip()
            officer_names = []
            for i in range(1, 7):
                officer = request.form.get(f'officer_names_{i}')
                if officer and officer.strip():
                    officer_names.append(officer.strip())
            activity = request.form['activity'].strip()

            # Validasi input
            if not activity:
                flash('Kegiatan harus diisi!', 'error')
                return redirect(url_for('add_document'))

            if not officer_names:
                flash('Pilih setidaknya satu petugas!', 'error')
                return redirect(url_for('add_document'))

            if len(officer_names) > 6:
                flash('Maksimal 6 petugas per kegiatan!', 'error')
                return redirect(url_for('add_document'))

            if not sequence_number.isdigit() or int(sequence_number) <= 0:
                flash('Nomor urut harus angka positif!', 'error')
                return redirect(url_for('add_document'))

            if not document_number:
                flash('Nomor surat harus diisi!', 'error')
                return redirect(url_for('add_document'))

            # Validasi nama petugas
            officer_names = list(dict.fromkeys(officer_names))  # Hilangkan duplikasi
            for officer in officer_names:
                if officer not in OFFICERS:
                    flash(f'Nama petugas {officer} tidak valid!', 'error')
                    return redirect(url_for('add_document'))

            # Parse tanggal
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = start_date if not end_date_str else datetime.strptime(end_date_str, '%Y-%m-%d').date()

            # Validasi rentang tanggal
            if end_date < start_date:
                flash('Tanggal selesai harus sama atau setelah tanggal mulai!', 'error')
                return redirect(url_for('add_document'))

            # Validasi hari Minggu
            date_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
            for date in date_range:
                if is_sunday(date):
                    flash(f'Tidak dapat menambahkan dokumen pada hari Minggu: {date.strftime("%Y-%m-%d")}!', 'error')
                    return redirect(url_for('add_document'))

            # Validasi tidak ada kegiatan lain di rentang
            if check_date_overlap(start_date, end_date):
                flash(f'Ada kegiatan lain dalam rentang tanggal {start_date.strftime("%Y-%m-%d")} - {end_date.strftime("%Y-%m-%d")}!', 'error')
                return redirect(url_for('add_document'))

            # Validasi petugas unik di rentang
            conflict, officer, conflict_date = check_officer_conflict(start_date, end_date, officer_names)
            if conflict:
                flash(f'Petugas {officer} sudah terdaftar di kegiatan lain pada tanggal {conflict_date.strftime("%Y-%m-%d")}!', 'error')
                return redirect(url_for('add_document'))

            # Validasi nomor surat unik
            if Document.query.filter_by(document_number=document_number).first():
                flash(f'Nomor surat {document_number} sudah digunakan!', 'error')
                return redirect(url_for('add_document'))

            # Buat dokumen baru
            new_doc = Document(
                sequence_number=int(sequence_number),
                start_date=start_date,
                end_date=end_date if end_date != start_date else None,
                officer_names=';'.join(officer_names),
                activity=activity,
                document_number=document_number
            )
            db.session.add(new_doc)
            db.session.commit()
            print(f"Added document: start_date={start_date}, end_date={end_date}, sequence_number={sequence_number}, document_number={document_number}")

            flash('Dokumen berhasil ditambahkan!', 'success')
            return redirect(url_for('index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan: {str(e)}', 'error')
            print(f"Error: {str(e)}")
            return redirect(url_for('add_document'))

    # Default document_number
    default_sequence = 1
    default_date = datetime.today().date()
    default_document_number = generate_document_number(default_sequence, default_date)
    return render_template('add.html', officers=OFFICERS, default_sequence=default_sequence, default_document_number=default_document_number)

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_document(id):
    document = Document.query.get_or_404(id)
    print(f"Editing document ID {id}: officer_names='{document.officer_names}'")

    if request.method == 'POST':
        try:
            start_date_str = request.form['start_date']
            end_date_str = request.form.get('end_date')
            sequence_number = request.form['sequence_number']
            document_number = request.form['document_number'].strip()
            officer_names = []
            for i in range(1, 7):
                officer = request.form.get(f'officer_names_{i}')
                if officer and officer.strip():
                    officer_names.append(officer.strip())
            activity = request.form['activity'].strip()

            # Validasi input
            if not activity:
                flash('Kegiatan harus diisi!', 'error')
                return redirect(url_for('edit_document', id=id))

            if not officer_names:
                flash('Pilih setidaknya satu petugas!', 'error')
                return redirect(url_for('edit_document', id=id))

            if len(officer_names) > 6:
                flash('Maksimal 6 petugas per kegiatan!', 'error')
                return redirect(url_for('edit_document', id=id))

            if not sequence_number.isdigit() or int(sequence_number) <= 0:
                flash('Nomor urut harus angka positif!', 'error')
                return redirect(url_for('edit_document', id=id))

            if not document_number:
                flash('Nomor surat harus diisi!', 'error')
                return redirect(url_for('edit_document', id=id))

            # Validasi nama petugas
            officer_names = list(dict.fromkeys(officer_names))
            for officer in officer_names:
                if officer not in OFFICERS:
                    flash(f'Nama petugas {officer} tidak valid!', 'error')
                    return redirect(url_for('edit_document', id=id))

            # Parse tanggal
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = start_date if not end_date_str else datetime.strptime(end_date_str, '%Y-%m-%d').date()

            # Validasi rentang tanggal
            if end_date < start_date:
                flash('Tanggal selesai harus sama atau setelah tanggal mulai!', 'error')
                return redirect(url_for('edit_document', id=id))

            # Validasi hari Minggu
            date_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
            for date in date_range:
                if is_sunday(date):
                    flash(f'Tidak dapat mengedit dokumen untuk hari Minggu: {date.strftime("%Y-%m-%d")}!', 'error')
                    return redirect(url_for('edit_document', id=id))

            # Validasi tidak ada kegiatan lain di rentang
            if check_date_overlap(start_date, end_date, exclude_id=id):
                flash(f'Ada kegiatan lain dalam rentang tanggal {start_date.strftime("%Y-%m-%d")} - {end_date.strftime("%Y-%m-%d")}!', 'error')
                return redirect(url_for('edit_document', id=id))

            # Validasi petugas unik di rentang
            conflict, officer, conflict_date = check_officer_conflict(start_date, end_date, officer_names, exclude_id=id)
            if conflict:
                flash(f'Petugas {officer} sudah terdaftar di kegiatan lain pada tanggal {conflict_date.strftime("%Y-%m-%d")}!', 'error')
                return redirect(url_for('edit_document', id=id))

            # Validasi nomor surat unik
            if Document.query.filter(Document.document_number == document_number, Document.id != id).first():
                flash(f'Nomor surat {document_number} sudah digunakan!', 'error')
                return redirect(url_for('edit_document', id=id))

            # Update dokumen
            document.sequence_number = int(sequence_number)
            document.start_date = start_date
            document.end_date = end_date if end_date != start_date else None
            document.officer_names = ';'.join(officer_names)
            document.activity = activity
            document.document_number = document_number

            db.session.commit()
            print(f"Updated document ID {id}: sequence_number={sequence_number}, document_number={document_number}")

            flash('Dokumen berhasil diperbarui!', 'success')
            return redirect(url_for('index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan: {str(e)}', 'error')
            print(f"Error: {str(e)}")
            return redirect(url_for('edit_document', id=id))

    return render_template('edit.html', document=document, officers=OFFICERS)

@app.route('/delete/<int:id>', methods=['POST'])
def delete_document(id):
    document = Document.query.get_or_404(id)
    try:
        print(f"Deleting document ID {id}: sequence_number={document.sequence_number}, document_number={document.document_number}")
        db.session.delete(document)
        db.session.commit()
        flash('Dokumen berhasil dihapus!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
        print(f"Error: {str(e)}")
    return redirect(url_for('index'))

@app.route('/download_excel')
def download_excel():
    documents = Document.query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riwayat Dokumen"

    headers = ['Tanggal', 'Nomor Urut', 'Nama Petugas', 'Kegiatan', 'Nomor Surat']
    ws.append(headers)

    # Format header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # Tulis data
    for row, doc in enumerate(documents, start=2):
        # Format officer_names dengan bullet dan newline
        officers = doc.officer_names.split(';')
        formatted_officers = '\n'.join([f"• {officer.strip()}" for officer in officers if officer.strip()])

        # Format tanggal sebagai rentang
        date_str = doc.start_date.strftime('%Y-%m-%d')
        if doc.end_date:
            date_str += f" - {doc.end_date.strftime('%Y-%m-%d')}"

        ws.append([
            date_str,
            doc.sequence_number,
            formatted_officers,
            doc.activity,
            doc.document_number
        ])

        # Aktifkan wrap text untuk kolom Nama Petugas
        cell = ws.cell(row=row, column=3)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    # Atur lebar kolom
    column_widths = {'A': 25, 'B': 10, 'C': 50, 'D': 40, 'E': 30}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Atur tinggi baris
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20 * len(ws.cell(row=row, column=3).value.split('\n'))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=riwayat_dokumen.xlsx'}
    )

with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True)